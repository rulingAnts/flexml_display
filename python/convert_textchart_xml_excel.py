#!/usr/bin/env python3
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from collections import Counter
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import logging
import os

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Placeholder shown when a <listRef> element has no text (user hasn't configured abbreviation)
LISTREF_EMPTY_PLACEHOLDER = '?'  # Change to e.g. 'Ø' or 'UNDEF' if preferred

# Clause type style map (fill only; text stays black)
CLAUSE_STYLE_MAP = {
    'normal':   {'fill': None},        # no shading for normal
    'dependent':{'fill': 'FFEDE9FE'},  # purple
    'speech':   {'fill': 'FFFEF3C7'},  # amber
    'song':     {'fill': 'FFFCE7F3'},  # rose
}

def clause_style_for_row_type(row_type: str | None) -> str:
    key = (row_type or '').lower()
    if key in CLAUSE_STYLE_MAP:
        return key
    # Fuzzy mapping for variant labels
    if 'depend' in key:
        return 'dependent'
    if 'speech' in key:
        return 'speech'
    if 'song' in key or 'poem' in key or 'hymn' in key:
        return 'song'
    # Treat unknown non-title rows as normal/independent
    return 'normal'

def parse_cell(cell, listref_counter: Counter | None = None, clause_targets: list | None = None):
    """Parse a chart cell.

    Output ordering rule (requested): gloss information (paired or grouped) must appear BEFORE any listRef code tokens.

    Strategy:
    1. Collect words and glosses first.
    2. Collect literal (non-parenthesis) tokens encountered in <main>.
    3. Collect listRef codes separately (do not interleave yet).
    4. Build a combined string: WORD/GLOSS content + literals + listRef tokens at the end.

    Mismatched counts (len(glosses) != len(words)) => emit words joined, then a single parenthetical group with all glosses.
    Matched counts => emit each word with its gloss: word (gloss).
        Clause markers:
            - Multiple <clauseMkr> elements are inspected. If they form a contiguous sequence (e.g. 19a,19b,19c) they are compressed to first-last (19a-19c).
            - If not contiguous (e.g. 19a,19c,20b) they are listed with commas: 19a,19c,20b.
            - A single marker appears alone: 19a.
            - The synthesized form is always enclosed in brackets: [19a-19c] or [19a,19c,20b].
            - Any literal bracket/dash characters found in the same cell that served the same purpose are suppressed to avoid duplication.
    """
    main = cell.find('main')
    if main is None:
        return ''

    # Row number (special first column)
    rownum = main.find('rownum')
    if rownum is not None:
        text = rownum.text or ''
        return text

    # Move marker
    move_mkr = main.find('moveMkr')
    if move_mkr is not None:
        text = move_mkr.text or ''
        return text

    # Note (for Notes column)
    note = main.find('note')
    if note is not None:
        text = note.text or ''
        return text

    # Pre-scan words and glosses for pairing
    words_all = [e.text.strip() for e in list(main) if e.tag == 'word' and e.text]
    glosses_elem = cell.find('glosses')
    gloss_texts = [g.text for g in glosses_elem.findall('gloss')] if glosses_elem is not None else []

    # Build a token stream from main children respecting <lit> spacing flags
    tokens: list[dict] = []  # each: {'text': str, 'type': str, 'nsb': bool, 'nsa': bool}
    word_index = 0
    for elem in list(main):
        tag = elem.tag
        text = (elem.text or '').strip() if elem.text else ''
        # Allow empty <lit>, <listRef> (so we can show placeholder), and <clauseMkr> (target attr)
        if not text and tag not in ('lit', 'listRef', 'clauseMkr'):
            continue
        if tag == 'lit':
            lit_text = (elem.text or '')
            nsb = (elem.get('noSpaceBefore') == 'true')
            nsa = (elem.get('noSpaceAfter') == 'true')
            tokens.append({'text': lit_text, 'type': 'literal', 'nsb': nsb, 'nsa': nsa})
        elif tag == 'word':
            tokens.append({'text': text, 'type': 'word', 'nsb': False, 'nsa': False})
            # If 1:1 pairing, add gloss immediately after word
            if gloss_texts and len(words_all) == len(gloss_texts) and word_index < len(gloss_texts):
                g = gloss_texts[word_index]
                if g:
                    tokens.append({'text': '(', 'type': 'literal', 'nsb': True, 'nsa': True})
                    tokens.append({'text': g, 'type': 'gloss', 'nsb': False, 'nsa': False})
                    tokens.append({'text': ')', 'type': 'literal', 'nsb': True, 'nsa': False})
            word_index += 1
        elif tag == 'listRef':
            # Prefer element text; fall back to common attributes like 'abbr' or 'code'
            code = text or elem.get('abbr') or elem.get('code') or LISTREF_EMPTY_PLACEHOLDER
            if listref_counter is not None and code:
                listref_counter[code] += 1
            tokens.append({'text': code, 'type': 'code', 'nsb': False, 'nsa': False})
        elif tag == 'clauseMkr':
            target = elem.get('target') or text or ''
            if target and clause_targets is not None:
                clause_targets.append(target)
            # Render clause marker as its text; brackets/dashes come from adjacent <lit>
            if target:
                tokens.append({'text': target, 'type': 'clause', 'nsb': False, 'nsa': False})

    # If glosses exist but not 1:1, append grouped gloss bundle at end
    if gloss_texts and len(words_all) != len(gloss_texts):
        joined = ' '.join(g for g in gloss_texts if g)
        if joined:
            tokens.append({'text': '(', 'type': 'literal', 'nsb': False, 'nsa': True})
            tokens.append({'text': joined, 'type': 'gloss', 'nsb': False, 'nsa': False})
            tokens.append({'text': ')', 'type': 'literal', 'nsb': True, 'nsa': False})

    # Now assemble string with spacing rules
    out_parts: list[str] = []
    prev_nsa = False
    for i, t in enumerate(tokens):
        txt = t['text']
        if not txt:
            continue
        nsb = t.get('nsb', False)
        nsa = t.get('nsa', False)
        # Decide spacing: insert space iff previous did not set noSpaceAfter and current did not set noSpaceBefore
        if out_parts:
            if not prev_nsa and not nsb:
                out_parts.append(' ')
        out_parts.append(txt)
        prev_nsa = nsa

    # Final string (no per-cell bracket balancing or listRef fallback)
    return ''.join(out_parts)

## Removed get_user_choice: script now always uses a single generic analysis column.

def convert_to_excel(input_file, output_file):
    analysis_choice = "Other/Custom"  # Forced default
    logging.debug("Starting conversion with input: %s, output: %s (forced analysis choice: %s)", input_file, output_file, analysis_choice)
    
    # Parse the XML file
    try:
        tree = ET.parse(input_file)
        root = tree.getroot()
        chart = root.find('chart')
    except ET.ParseError as e:
        logging.error("XML parsing failed: %s", str(e))
        raise Exception(f"Failed to parse XML file: {str(e)}")

    # Extract headers
    title1_row = chart.find('row[@type="title1"]')
    title2_row = chart.find('row[@type="title2"]')

    # Dynamically extract main headers and their spans from title1 row
    main_headers = [('', 1)]  # Blank column
    for cell in title1_row.findall('cell'):
        cols = int(cell.get('cols', 1))
        lit = cell.find('main/lit')
        header_text = lit.text if lit is not None else ''
        main_headers.append((header_text, cols))

    # Always use the simplified generic analysis placeholder
    analysis_columns = ["Add Analysis Columns Here"]
    main_headers.insert(-1, ('Add Analysis Columns Here', 1))

    # Extract sub-headers from title2 row, starting from cell[1] to cell[9]
    sub_headers = ['', 'Row']  # Blank column and Row
    for cell in title2_row.findall('cell')[1:-1]:  # From Outer (Topic) to Outer
        lit = cell.find('main/lit')
        sub_headers.append(lit.text if lit is not None else '')
    sub_headers.extend(analysis_columns)  # Add analysis columns
    sub_headers.append('Notes')  # Notes

    # Verify the total number of columns matches
    total_cols = sum(span for _, span in main_headers)
    if total_cols != len(sub_headers):
        logging.error("Header column mismatch: %d in title1, %d in title2", total_cols, len(sub_headers))
        raise Exception(f"Header column mismatch: {total_cols} columns in title1 (including blank and new), {len(sub_headers)} in title2")

    # Collect listRef codes across the chart for legend + include in parsing
    listref_counter = Counter()

    # Precompute mapping of row id -> type (to color clause markers by target type)
    row_type_by_id: dict[str, str] = {}
    for r in chart.findall('row'):
        rtype = r.get('type')
        if rtype in ('title1', 'title2'):
            continue
        rid = r.get('id') or ''
        if rid:
            row_type_by_id[rid] = rtype or 'normal'

    # Extract data rows
    data = []
    merges_by_row: list[list[tuple[int, int]]] = []  # per data row: list of (start_col, end_col) to merge
    break_flags: list[str | None] = []  # 'sent', 'para', or None per data row order
    row_ids_in_order: list[str] = []
    rownums_in_order: list[str] = []
    row_types_in_order: list[str] = []
    clause_target_global: list[str] = []
    for row in chart.findall('row'):
        rtype = row.get('type')
        if rtype in ('title1', 'title2'):
            continue
        row_data = ['']  # Blank leading column
        row_cells = row.findall('cell')
        if not row_cells:
            continue
        body_cells = row_cells[:-1]
        notes_cell = row_cells[-1]
        rownum_text_for_row = ''
        style_key = clause_style_for_row_type(rtype)
        row_merges: list[tuple[int, int]] = []
        current_col = 2  # Column B is sub_headers[1] ('Row') since col 1 is blank
        for cell in body_cells:
            span = int(cell.get('cols', 1))
            local_clause_targets: list[str] = []
            # Capture rownum text if present in this cell
            try:
                rn = cell.find('main/rownum')
                if rn is not None and rn.text:
                    rownum_text_for_row = (rn.text or '').strip()
            except Exception:
                pass
            value = parse_cell(cell, listref_counter, local_clause_targets)
            if local_clause_targets:
                clause_target_global.extend(local_clause_targets)
            row_data.append(value)
            if span > 1:
                # Record merge across these columns for this row
                start_c = current_col
                end_c = current_col + span - 1
                row_merges.append((start_c, end_c))
            if span > 1:
                for _ in range(span - 1):
                    row_data.append('')
            current_col += span
        # Analysis placeholder columns (no runs)
        row_data.extend([''] * len(analysis_columns))
        # Notes
        note_val = parse_cell(notes_cell, listref_counter)
        row_data.append(note_val)
        # Normalize length
        if len(row_data) > len(sub_headers):
            logging.warning("Row %s produced %d columns; trimming to %d", row.get('id'), len(row_data), len(sub_headers))
            row_data = row_data[:len(sub_headers)]
        elif len(row_data) < len(sub_headers):
            pad = len(sub_headers) - len(row_data)
            logging.warning("Row %s produced %d columns; padding to %d", row.get('id'), len(row_data), len(sub_headers))
            row_data.extend([''] * pad)
        data.append(row_data)
        merges_by_row.append(row_merges)
        # Determine break type for this row
        if row.get('endPara') == 'true':
            break_flags.append('para')
        elif row.get('endSent') == 'true':
            break_flags.append('sent')
        else:
            break_flags.append(None)
        row_ids_in_order.append(row.get('id') or '')
        rownums_in_order.append(rownum_text_for_row)
        row_types_in_order.append(rtype or 'normal')

    # Create DataFrame without headers
    try:
        df = pd.DataFrame(data, columns=sub_headers)
    except ValueError as e:
        logging.error("DataFrame creation failed: %s", str(e))
        raise Exception(f"DataFrame creation failed: {str(e)}")

    # Write to Excel, starting data at row 3 (will become row 7 after inserting 4 rows)
    df.to_excel(output_file, index=False, header=False, startrow=2, engine='openpyxl')

    # Load the workbook to apply formatting
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Insert four rows at the top for blank, title, subtitle, and blank
    ws.insert_rows(1, amount=4)

    # Define border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin')
    )
    thick_border_left = Border(left=Side(style='thick'))
    thick_border_right = Border(right=Side(style='thick'))
    thick_border_top = Border(top=Side(style='thick'))
    thick_border_bottom = Border(bottom=Side(style='thick'))

    # Add blank row (row 1)
    for col_idx in range(1, len(sub_headers) + 1):
        ws.cell(row=1, column=col_idx).value = None

    # Add title row (row 2)
    title_cell = ws.cell(row=2, column=2)  # Start at column B
    title_cell.value = "Story Title"
    title_cell.font = Font(bold=True, size=16)  # Bold, larger font for title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(sub_headers))  # Merge B to last

    # Add subtitle row (row 3)
    subtitle_cell = ws.cell(row=3, column=2)  # Start at column B
    subtitle_cell.value = "Text Chart"
    subtitle_cell.font = Font(bold=True, size=14)  # Bold, slightly smaller than title
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(sub_headers))  # Merge B to last

    # Add blank row (row 4)
    for col_idx in range(1, len(sub_headers) + 1):
        ws.cell(row=4, column=col_idx).value = None

    # Write main headers (row 5) and merge cells
    col_offset = 2  # Start after blank column
    for header, span in main_headers[1:]:  # Skip blank column header
        if header:
            cell = ws.cell(row=5, column=col_offset)
            cell.value = header
            cell.font = Font(bold=True, size=14)  # Bold, 2 points larger than default (12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if span > 1:
                ws.merge_cells(start_row=5, start_column=col_offset, end_row=5, end_column=col_offset + span - 1)
        col_offset += span

    # Extend Notes header to span two rows
    notes_col = col_offset - 1
    ws.cell(row=5, column=notes_col).value = 'Notes'
    ws.cell(row=5, column=notes_col).font = Font(bold=True, size=14)
    ws.cell(row=5, column=notes_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=5, start_column=notes_col, end_row=6, end_column=notes_col)

    # Write sub-headers (row 6), skipping Notes column
    for col_idx, header in enumerate(sub_headers[:-1], start=1):  # Exclude Notes (merged)
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)  # Bold, default size
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge data row cells according to cols attribute (like HTML colspan)
    # Data rows begin at Excel row 7, columns B..last
    for idx, row_merges in enumerate(merges_by_row):
        excel_row = 7 + idx
        for (start_c, end_c) in row_merges:
            # Guard against out-of-range merges
            if start_c < 2 or end_c > ws.max_column or end_c <= start_c:
                continue
            try:
                ws.merge_cells(start_row=excel_row, start_column=start_c, end_row=excel_row, end_column=end_c)
            except Exception as e:
                logging.debug("Merge failed at row %d, cols %d-%d: %s", excel_row, start_c, end_c, e)

    # Apply subtle (thin) borders to all cells within the chart area (rows 5..max_row, cols B..last)
    thin_all = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row_idx in range(5, ws.max_row + 1):
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_all
            # For readability, center content in merged header/data when applicable; default left otherwise
            if row_idx in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Re-apply thick box border around chart (rows 5 to max_row, columns B to last) to override thin edges
    # Top border (row 5, columns B to last)
    for col_idx in range(2, ws.max_column + 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.border = Border(
            top=Side(style='thick'),
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

    # Bottom border (last data row, columns B to last)
    for col_idx in range(2, ws.max_column + 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.border = Border(
            bottom=Side(style='thick'),
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top
        )

    # Left border (column B, rows 5 to max_row)
    for row_idx in range(5, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=2)
        cell.border = Border(
            left=Side(style='thick'),
            right=cell.border.right,
            top=cell.border.top,
            bottom=cell.border.bottom
        )

    # Right border (last column, rows 5 to max_row)
    for row_idx in range(5, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=ws.max_column)
        cell.border = Border(
            right=Side(style='thick'),
            left=cell.border.left,
            top=cell.border.top,
            bottom=cell.border.bottom
        )

    # Apply sentence / paragraph break bottom borders inside the chart
    # Data rows start at Excel row 7; iterate over break_flags aligned to data order
    for idx, flag in enumerate(break_flags):
        if not flag:
            continue
        excel_row = 7 + idx
        # Choose border style
        if flag == 'para':
            bottom_side = Side(style='thick')
        else:  # 'sent'
            # Upgraded from thin to medium for greater visual separation
            bottom_side = Side(style='medium', color='000000')
        for col_idx in range(2, ws.max_column + 1):  # Columns B .. last (exclude leading blank col A)
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=bottom_side
            )

    # (Removed conditional data validation; not relevant for single generic analysis column.)

    # Rich text styling removed; all text remains plain black.

    # Clause row shading (moved outside rich text block so it always runs)
    shaded_rows_count = 0
    # Shade all non-normal rows by their type (independent of clause targets)
    for idx, rtype in enumerate(row_types_in_order):
        style_key = clause_style_for_row_type(rtype)
        fill_color = CLAUSE_STYLE_MAP.get(style_key, CLAUSE_STYLE_MAP['normal'])['fill']
        if not fill_color:
            continue
        erow = 7 + idx
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for cidx in range(2, ws.max_column + 1):
            cell = ws.cell(row=erow, column=cidx)
            cell.fill = fill
        shaded_rows_count += 1
    logging.debug("Colored clause markers: %d | Shaded rows: %d", 0, shaded_rows_count)

    # Text color disabled: skip safe coloring pass

    # Adjust column widths, handling merged cells
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue  # Skip merged cells
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # (Temporarily disable auto-filter to rule out hidden rows due to filters)
    # ws.auto_filter.ref = f"B6:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Add Debug sheet with row id/rownum mapping and clause target markers
    try:
        if 'Debug' in wb.sheetnames:
            dbg = wb['Debug']
            for row in dbg['A1:Z999']:
                for cell in row:
                    cell.value = None
        else:
            dbg = wb.create_sheet(title='Debug')
        dbg['A1'] = 'ExcelRow'
        dbg['B1'] = 'RowId'
        dbg['C1'] = 'Rownum'
        dbg['D1'] = 'ClauseTargeted'
        dbg['A1'].font = Font(bold=True)
        dbg['B1'].font = Font(bold=True)
        dbg['C1'].font = Font(bold=True)
        dbg['D1'].font = Font(bold=True)
        targets_set = set(clause_target_global or [])
        for idx, (rid, rnum) in enumerate(zip(row_ids_in_order, rownums_in_order), start=0):
            erow = 7 + idx
            dbg.cell(row=2 + idx, column=1, value=erow)
            dbg.cell(row=2 + idx, column=2, value=rid)
            dbg.cell(row=2 + idx, column=3, value=rnum)
            # Mark if matched by exact target string
            dbg.cell(row=2 + idx, column=4, value='Y' if (rid in targets_set or rnum in targets_set) else '')
        dbg.column_dimensions['A'].width = 10
        dbg.column_dimensions['B'].width = 20
        dbg.column_dimensions['C'].width = 12
        dbg.column_dimensions['D'].width = 14
    except Exception as e:
        logging.debug("Failed to add Debug sheet: %s", e)

    # Add Legend sheet for listRef codes and clause type colors
    try:
        if 'Legend' in wb.sheetnames:
            legend_ws = wb['Legend']
            # Clear existing (simple approach)
            for row in legend_ws['A1:Z999']:
                for cell in row:
                    cell.value = None
        else:
            legend_ws = wb.create_sheet(title='Legend')
        # Section 1: listRef codes
        legend_ws['A1'] = 'Code'
        legend_ws['B1'] = 'Count'
        legend_ws['A1'].font = Font(bold=True)
        legend_ws['B1'].font = Font(bold=True)
        for idx, (code, count) in enumerate(sorted(listref_counter.items()), start=2):
            legend_ws.cell(row=idx, column=1, value=code)
            legend_ws.cell(row=idx, column=2, value=count)
        listref_rows = max(1, len(listref_counter))
        cursor = 2 + listref_rows + 1  # one blank line after listRef section

        # Section 2: Clause Types color key
        legend_ws.cell(row=cursor, column=1, value='Clause Types').font = Font(bold=True)
        cursor += 1
        legend_ws.cell(row=cursor, column=1, value='Type').font = Font(bold=True)
        legend_ws.cell(row=cursor, column=2, value='Marker Text').font = Font(bold=True)
        legend_ws.cell(row=cursor, column=3, value='Row Shading').font = Font(bold=True)
        cursor += 1
        # Order types for readability
        types_order = ['dependent', 'speech', 'song', 'normal']
        for tkey in types_order:
            style = CLAUSE_STYLE_MAP.get(tkey, CLAUSE_STYLE_MAP['normal'])
            name = tkey.capitalize()
            legend_ws.cell(row=cursor, column=1, value=name)
            # Sample marker text with font color
            legend_ws.cell(row=cursor, column=2, value='[19a-19b]')
            # Fill swatch (or none)
            fill_color = style.get('fill')
            if fill_color:
                legend_ws.cell(row=cursor, column=3, value='').fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            else:
                legend_ws.cell(row=cursor, column=3, value='(none)')
            cursor += 1

        # Adjust widths
        legend_ws.column_dimensions['A'].width = 18
        legend_ws.column_dimensions['B'].width = 16
        legend_ws.column_dimensions['C'].width = 14
    except Exception as e:
        logging.warning("Failed to add Legend sheet: %s", e)

    # Save the formatted workbook
    wb.save(output_file)
    logging.debug("Excel file saved: %s", output_file)

def main():
    # Initialize Tkinter
    root = tk.Tk()
    root.withdraw()  # Hide the main Tkinter window

    try:
        # No analysis choice prompt (simplified UX)

        # Open file dialog for input file
        logging.debug("Opening file dialog for input")
        input_file = filedialog.askopenfilename(
            title="Select Input XML File",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")]
        )
        if not input_file:
            logging.info("No input file selected. Exiting.")
            messagebox.showinfo("Info", "No input file selected. Exiting.")
            return

        # Open file dialog for output file with overwrite confirmation
        while True:
            logging.debug("Opening file dialog for output")
            output_file = filedialog.asksaveasfilename(
                title="Save Output Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not output_file:
                logging.info("No output file selected. Exiting.")
                messagebox.showinfo("Info", "No output file selected. Exiting.")
                return
            if os.path.exists(output_file):
                logging.debug("Selected output file exists: %s", output_file)
                overwrite = messagebox.askyesno(
                    "Confirm Overwrite",
                    f"The file:\n{output_file}\nalready exists. Overwrite?"
                )
                if overwrite:
                    break
                else:
                    logging.debug("User declined overwrite; re-prompting for filename")
                    continue  # Loop back for a new filename
            else:
                break  # File does not exist; safe to proceed

        # Convert the file
        logging.debug("Starting conversion")
        convert_to_excel(input_file, output_file)
        messagebox.showinfo("Success", f"Excel file '{output_file}' has been created successfully.")

    except Exception as e:
        logging.error("Error occurred: %s", str(e))
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

    finally:
        # The root.destroy() should be handled by the application closing after the mainloop
        # For now, we'll keep it here, but it's part of the problem.
        # A more robust solution would manage the root window lifespan differently.
        root.destroy()
        logging.debug("Tkinter root destroyed")

if __name__ == "__main__":
    main()